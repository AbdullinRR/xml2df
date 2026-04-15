import tempfile
import os

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from streamlit_tree_select import tree_select

from xml2df import XML2DF

SHEET_NAME = "Данные"
HEADER_COLOR = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"
FONT_NAME = "Arial"
FONT_SIZE = 10
MAX_COL_WIDTH = 60


def save_df_to_xlsx(df: pd.DataFrame, output_path: str) -> None:
    """
    Сохраняет DataFrame в xlsx с форматированием.

    Args:
        df: DataFrame с данными.
        output_path: Путь к выходному файлу.
    """
    df.to_excel(output_path, index=False, sheet_name=SHEET_NAME)

    wb = load_workbook(output_path)
    ws = wb[SHEET_NAME]

    header_fill = PatternFill("solid", start_color=HEADER_COLOR)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, name=FONT_NAME, size=FONT_SIZE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # форматируем шапку
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    # форматируем строки данных
    data_font = Font(name=FONT_NAME, size=FONT_SIZE)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(vertical="top")

    # подбираем ширину колонок
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, MAX_COL_WIDTH)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


st.title("XML → XLSX")

uploaded_file = st.file_uploader("Загрузи XML файл", type=["xml"])

if uploaded_file is not None:
    with tempfile.NamedTemporaryFile(suffix=".xml", delete=False) as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

    try:
        converter = XML2DF()
        tree_nodes = converter.get_tree_nodes(tmp_path)

        st.subheader("Выбери теги для конвертации")
        selected = tree_select(
            tree_nodes,
            check_model="leaf",
            show_expand_all=True,
        )

        checked = selected.get("checked", [])

        if checked and st.button("Конвертировать"):
            try:
                with tempfile.TemporaryDirectory() as tmp_dir:
                    xlsx_path = os.path.join(tmp_dir, "output.xlsx")

                    df = converter.get_df_from_xml(tmp_path)
                    df = df.loc[:, ~df.columns.duplicated()]

                    # строим маппинг: колонка df → полный путь из checked
                    col_rename = {}
                    for col in df.columns:
                        col_parts = col.split("|")
                        best_match = None
                        best_len = -1
                        for checked_path in checked:
                            checked_parts = checked_path.split("|")
                            if checked_parts[-len(col_parts):] == col_parts and len(checked_parts) > best_len:
                                best_match = checked_path
                                best_len = len(checked_parts)
                        if best_match and col not in col_rename:
                            col_rename[col] = best_match

                    # переименовываем и убираем дубликаты
                    df = df.rename(columns=col_rename)
                    df = df.loc[:, ~df.columns.duplicated()]

                    # нормализуем checked — убираем #value суффикс
                    normalized_checked = []
                    for c in checked:
                        if c.endswith("#value"):
                            normalized_checked.append(c.replace("#value", ""))
                        else:
                            normalized_checked.append(c)

                    # убираем дубликаты сохраняя порядок
                    seen = set()
                    normalized_checked = [x for x in normalized_checked if not (x in seen or seen.add(x))]

                    selected_cols = [c for c in normalized_checked if c in df.columns]
                    df = df[selected_cols]

                    save_df_to_xlsx(df, xlsx_path)

                    with open(xlsx_path, "rb") as f:
                        xlsx_bytes = f.read()

                st.subheader("DataFrame")
                st.dataframe(df)
                st.write(f"Строк: {len(df)}, столбцов: {len(df.columns)}")

                st.download_button(
                    label="Скачать XLSX",
                    data=xlsx_bytes,
                    file_name=uploaded_file.name.replace(".xml", ".xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            except Exception as e:
                st.error(f"Ошибка: {e}")

    except Exception as e:
        st.error(f"Ошибка чтения XML: {e}")
    finally:
        os.unlink(tmp_path)